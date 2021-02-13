import requests
from bs4 import BeautifulSoup
import pyautogui
import os
import openpyxl

keyword = pyautogui.prompt(text="검색어를 입력하세요", title="Message")
last_page = int(pyautogui.prompt(text="몇 페이지까지 검색할까요?", title="Message"))

cur_page = 1 # 현재 페이지 

# 리스트 생성
news_title_list = [] # 뉴스 제목 리스트
news_content_list = [] # 뉴스 본문 리스트

for i in range(1, last_page*10 + 1, 10): # 1, 11, 21, .... last_page*10 + 1

    print(f'========현재 페이지 {cur_page}=========')
    url = f'https://search.naver.com/search.naver?&where=news&query={keyword}&start={i}'

    response = requests.get(url)
    response.raise_for_status()
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    headers = soup.select("div.info_group")

    for header in headers:
        links = header.select("a.info")

        if len(links) > 1:
            # 네이버 뉴스 링크 url 가져오기 links[0] : 언론사 홈페이지 링크
            sub_url = links[1].attrs['href']
            
            # 예외상황
            # 비정상적인 접근이라고 판단 (Connection aborted) -> 사용자 인 것처럼 (브라우저에서 접근한 것처럼)
            response = requests.get(sub_url, headers={'User-Agent' : 'Mozilla/5.0'})
            response.raise_for_status()
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')

            title = soup.select_one("#articleTitle").get_text(strip=True) # 뉴스 제목(데이터)
            content = soup.select_one("#articleBodyContents").get_text(strip=True) #뉴스 본문(데이터)

            print(title)
            print(content)

            news_title_list.append(title)
            news_content_list.append(content)
    
    cur_page = cur_page + 1

print("크롤링 완료!! 엑셀에 저장합니다...")

# 저장 경로
save_path = r"C:\Users\스타트코딩\Desktop\웹크롤링\6주차\결과.xlsx"

# 엑셀 생성 (파일이 없으면 만들고, 있으면 만들지 않는다)
if not os.path.exists(save_path):
    openpyxl.Workbook().save(save_path)

# 엑셀 불러오기
workbook = openpyxl.load_workbook(save_path)

# 시트 생성
sheet = workbook.create_sheet(keyword)

# 데이터 저장하기
for i in range(len(news_title_list)):
    sheet[f'A{i+1}'] = news_title_list[i]
    sheet[f'B{i+1}'] = news_content_list[i]

# 엑셀 저장
workbook.save(save_path)