import requests
from bs4 import BeautifulSoup
import pyautogui
import openpyxl
import os

keyword = pyautogui.prompt(text='검색어를 입력하세요', title='Message')
page = int(pyautogui.prompt(text='몇 페이지까지 크롤링 할까요?', title='Message'))

# 저장 경로
save_path = r"C:\Users\스타트코딩\Desktop\웹크롤링\5주차\결과.xlsx"

# 엑셀 생성 (파일이 없으면 만들고, 있으면 만들지 않는다)
if not os.path.exists(save_path):
    openpyxl.Workbook().save(save_path)

# 엑셀 불러오기
workbook = openpyxl.load_workbook(save_path)

# 시트 생성
sheet = workbook.create_sheet(keyword)

for i in range(1, page + 1):
    url = f"https://kin.naver.com/search/list.nhn?query={keyword}&page={i}"

    print(f'=================={i}번째 페이지 입니다.====================')
    response = requests.get(url)
    response.raise_for_status()
    html = response.text

    soup = BeautifulSoup(html, 'html.parser')

    lists = soup.select('ul.basic1 > li')

    # 현재 행 번호
    row_num = 1 + (i-1) * 10

    for li in lists:
        link_url = li.select_one('dl > dt > a').attrs['href']

        response_new = requests.get(link_url)
        response_new.raise_for_status()
        html_new = response_new.text
        soup_new = BeautifulSoup(html_new, 'html.parser')
        title = soup_new.select_one("#content div.c-heading__title-inner > div.title").get_text(strip=True)
        content = soup_new.select("#content div.c-heading__content")
        
        # 있으면 [content의 html객체]
        # 없으면 [] 빈리스트
        
        if len(content) == 1:
            content = content[0].get_text(strip=True)
        else:
            content = "없음"

        sheet[f'A{row_num}'] = title
        sheet[f'B{row_num}'] = content

        workbook.save(save_path)
        row_num = row_num + 1
